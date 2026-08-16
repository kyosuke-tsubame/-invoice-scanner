#!/usr/bin/env python3
"""
納品書の記帳をExcelファイル（OneDrive）で管理する。
Googleスプレッドシート連携（curlがGoogle側にブロックされる問題があった）を廃止し、
FAX仕分けの売上集計.xlsxと同じ考え方（openpyxlでローカルのExcelファイルを直接読み書き）に統一した。

使い方:
  python3 invoice_ledger.py check '{"date":"2026-08-15","store":"本店","total":3950}'
    -> 同じ日付・店舗・金額の行が既にあれば {"duplicate": true} を返す

  python3 invoice_ledger.py add '{"date":"2026-08-15","store":"本店","supplier":"株式会社マルフク","total":3950,"category":"仕入れ"}'
    -> 「データ」シートに1行追記し、「店舗別集計」シートを最新の内容で作り直す

シート構成:
  データ     … 生ログ（追記のみ、日付/店舗/仕入先/金額/年月/区分）
  店舗別集計 … 店舗ごとに区切られた、仕入先×月の金額一覧（addのたびに全体を作り直す）
"""
import sys
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX_PATH = Path.home() / "Library/CloudStorage/OneDrive-個人用/納品書集計/納品書集計.xlsx"
DATA_SHEET = "データ"
SUMMARY_SHEET = "店舗別集計"
HEADER = ["日付", "店舗", "仕入先", "金額", "年月", "区分"]

STORE_ORDER = ["本店", "KADODE店", "空港店", "静岡紺屋町店", "セントラル", "冷凍事業部", "製麺事業部"]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1C2340", end_color="1C2340", fill_type="solid")
STORE_FONT = Font(bold=True, size=13)
SUBTOTAL_FONT = Font(bold=True)
MONEY_FORMAT = "#,##0"


def _load_workbook():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.is_file():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    if DATA_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(DATA_SHEET, 0)
        ws.append(HEADER)
        for col, _ in enumerate(HEADER, start=1):
            ws.cell(row=1, column=col).font = HEADER_FONT
            ws.cell(row=1, column=col).fill = HEADER_FILL
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
    return wb


def _read_rows(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "date": str(row[0]),
            "store": row[1],
            "supplier": row[2],
            "total": int(row[3] or 0),
            "year_month": row[4],
            "category": row[5],
        })
    return rows


def cmd_check(payload):
    wb = _load_workbook()
    ws = wb[DATA_SHEET]
    date, store, total = payload["date"], payload["store"], int(payload["total"])
    dup = any(r["date"] == date and r["store"] == store and r["total"] == total for r in _read_rows(ws))
    print(json.dumps({"duplicate": dup}, ensure_ascii=False))


def cmd_add(payload):
    wb = _load_workbook()
    ws = wb[DATA_SHEET]
    date = payload["date"]
    store = payload["store"]
    supplier = payload.get("supplier", "")
    total = int(payload["total"])
    category = payload.get("category", "")
    year_month = date[:7] if date else ""

    ws.append([date, store, supplier, total, year_month, category])
    ws.cell(row=ws.max_row, column=4).number_format = MONEY_FORMAT

    _rebuild_summary(wb, _read_rows(ws))
    wb.save(XLSX_PATH)
    print(f"追加: {date} {store} {supplier} {total}円")


def _rebuild_summary(wb, rows):
    if SUMMARY_SHEET in wb.sheetnames:
        wb.remove(wb[SUMMARY_SHEET])
    ws = wb.create_sheet(SUMMARY_SHEET)

    # 店舗 -> 仕入先 -> 年月 -> 金額合計
    agg = {}
    months = set()
    for r in rows:
        store_agg = agg.setdefault(r["store"], {})
        supplier_agg = store_agg.setdefault(r["supplier"] or "（仕入先不明）", {})
        supplier_agg[r["year_month"]] = supplier_agg.get(r["year_month"], 0) + r["total"]
        if r["year_month"]:
            months.add(r["year_month"])
    month_cols = sorted(months)

    # データに登場する順（STORE_ORDERに無い店舗名にも対応）
    stores_in_data = [s for s in STORE_ORDER if s in agg] + [s for s in agg if s not in STORE_ORDER]

    r_idx = 1
    max_col = max(2 + len(month_cols), 3)
    for store in stores_in_data:
        ws.cell(row=r_idx, column=1, value=store).font = STORE_FONT
        r_idx += 1

        header = ["仕入先"] + month_cols + ["合計"]
        for c_idx, label in enumerate(header, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        r_idx += 1

        supplier_totals = agg[store]
        store_total = 0
        month_totals = {m: 0 for m in month_cols}
        for supplier, by_month in sorted(supplier_totals.items(), key=lambda kv: -sum(kv[1].values())):
            row_total = sum(by_month.values())
            store_total += row_total
            ws.cell(row=r_idx, column=1, value=supplier)
            for c_idx, m in enumerate(month_cols, start=2):
                v = by_month.get(m, 0)
                month_totals[m] += v
                cell = ws.cell(row=r_idx, column=c_idx, value=v if v else None)
                cell.number_format = MONEY_FORMAT
            total_cell = ws.cell(row=r_idx, column=2 + len(month_cols), value=row_total)
            total_cell.number_format = MONEY_FORMAT
            r_idx += 1

        # 店舗の小計行
        ws.cell(row=r_idx, column=1, value="小計").font = SUBTOTAL_FONT
        for c_idx, m in enumerate(month_cols, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=month_totals[m] if month_totals[m] else None)
            cell.number_format = MONEY_FORMAT
            cell.font = SUBTOTAL_FONT
        total_cell = ws.cell(row=r_idx, column=2 + len(month_cols), value=store_total)
        total_cell.number_format = MONEY_FORMAT
        total_cell.font = SUBTOTAL_FONT
        r_idx += 2  # 空行を挟んで次の店舗へ

    ws.column_dimensions["A"].width = 26
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


def main():
    if len(sys.argv) != 3 or sys.argv[1] not in ("check", "add"):
        print("usage: invoice_ledger.py <check|add> '<JSON payload>'", file=sys.stderr)
        sys.exit(1)
    payload = json.loads(sys.argv[2])
    if sys.argv[1] == "check":
        cmd_check(payload)
    else:
        cmd_add(payload)


if __name__ == "__main__":
    main()
